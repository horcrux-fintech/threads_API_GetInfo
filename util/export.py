from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone

def export_excel(rows: list[dict], out_path: str):
    """
    將資料匯出為 Excel 檔案
    
    參數：
        rows (list[dict]): 
            含多筆貼文資料的清單，每筆為一個 dict，例如：
            [
                {
                    "id": "123456789",
                    "created_time": "2025-11-01T12:00:00+0000",
                    "text": "這是一篇貼文",
                    "views": 120,
                    "likes": 45,
                    "replies": 10,
                    "reposts": 5,
                    "quotes": 2,
                    "shares": 3,
                    "permalink": "https://www.threads.net/@user/post/123456"
                },
                ...
            ]
        out_path (str):
            輸出的檔案名稱與路徑，例如 "threads_posts_20251103.xlsx"。
    """

    fieldnames = [
        ("id", "貼文ID"),
        ("created_time", "發布時間"),
        ("text", "內文"),
        ("views", "觀看數"),
        ("likes", "按讚數"),
        ("replies", "回覆數"),
        ("reposts", "轉發數"),
        ("quotes", "引用數"),
        ("shares", "分享數"),
        ("permalink", "貼文連結")
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Threads 貼文"

    # 表頭
    headers = [cn for _, cn in fieldnames]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    # 資料
    for r in rows:
        ws.append([r.get(en, "") for en, _ in fieldnames])

    # 文字換行（text 欄）
    text_col_idx = [i for i, (en, _) in enumerate(fieldnames, start=1) if en == "text"]
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if text_col_idx and col_idx == text_col_idx[0]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 簡易自動欄寬
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            ln = len(str(v)) if v is not None else 0
            max_len = ln if ln > max_len else max_len
        ws.column_dimensions[letter].width = min(max_len * 1.6 + 2, 80)

    wb.save(out_path)
    print(f"Excel 輸出完成：{ws.max_row-1} 筆 -> {out_path}")

def export_tw_time(utc_str: str) -> str:
    """
    將 Threads API 回傳的 UTC 時間（例如 2025-11-06T10:23:05+0000）
    轉換為台灣時間（UTC+8），並回傳格式化字串
    """
    try:
        dt_utc = datetime.strptime(utc_str, "%Y-%m-%dT%H:%M:%S%z")
        dt_tw = dt_utc.astimezone(timezone(timedelta(hours=8)))
        return dt_tw.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return utc_str 